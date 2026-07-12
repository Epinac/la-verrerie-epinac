"""
Génère QUESTIONNAIRE_MAMAN.xlsx — VERSION 2 simplifiée pour ouverture V1.

Format : Section | N° | Question | Aide/exemple | 💡 Conseil de Morgan | Ta réponse

V2 (simplification post-feedback) :
- Plus de questions juridiques (statut, SIRET, TVA, capital) — pas pour le moment
- Adresse + GPS pré-remplis
- Section mariages / séminaires allégée pour V1
- Section légal simplifiée
- Section photos retirée (juste note d'intention pour le logo)
- Section témoignages retirée (pas encore de clients)
- Section carnet d'adresses simplifiée
- Section pinterest/moodboard retirée
- Section infrastructure ajoutée (fibre, électricité, mobilier événementiel)
- Aide à la décision sur le nom du lieu
- Question salle séminaire VS chapiteau extérieur
- Mention "si tu ne sais pas, on cherchera ensemble" pour les prix
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

ROOT = Path("/Users/morganspirli/Desktop/La Verrerie")
out = ROOT / "QUESTIONNAIRE_MAMAN.xlsx"

# (section, num, question, aide, conseil, type, prefill)
# type: "text" / "long" / "choice" / "info"
# prefill: réponse pré-remplie par Morgan (optionnel, défaut "")
QUESTIONS = [
    ("0. Mode d'emploi", "", "📖 LIS-MOI EN PREMIER",
     "Aucune pression. Mets le maximum de ce que tu sais. Si tu ne sais pas, écris « À voir ensemble ». Si pas applicable, écris « Sans objet ». Ne remplis QUE la colonne « Ta réponse ». Tu peux faire en plusieurs séances. 80% rempli > 100% parfait. Pour les prix : si tu ne sais pas, écris « À voir ensemble » — on cherchera ensemble.",
     "", "info", ""),

    # ===== 1. NOM DE LA MAISON =====
    ("1. Le nom de la maison", "1.1", "As-tu déjà choisi un nom pour la maison ?",
     "Oui (lequel) / J'hésite entre plusieurs / Pas encore d'idée.",
     "Si tu as plusieurs idées, écris-les toutes. On en choisira une ensemble.",
     "text", ""),
    ("1. Le nom de la maison", "1.2", "Si oui, quel nom ?",
     "Le nom officiel qui apparaîtra partout.",
     "Choisis un nom court, mémorable, qu'on peut épeler sans ambiguïté.",
     "text", ""),
    ("1. Le nom de la maison", "1.3", "Si tu hésites — 3 mots qui décrivent l'esprit du lieu",
     "Ex : « calme, ancien, lumineux » ou « familial, terre-à-terre, bourgogne ».",
     "Ces mots aideront à proposer des noms qui sonnent juste.",
     "text", ""),
    ("1. Le nom de la maison", "1.4", "Y a-t-il un élément historique fort qui pourrait inspirer le nom ?",
     "Un ancien usage du lieu, une famille, un événement, un métier.",
     "Ex : 'La Verrerie' parce que ancienne demeure des directeurs de la verrerie.",
     "long", ""),
    ("1. Le nom de la maison", "1.5", "Y a-t-il un détail du lieu qui pourrait inspirer le nom ?",
     "Un arbre du parc, une tour, une porte, un puits, une plante…",
     "Ex : 'Le Tilleul', 'La Tour Carrée', 'Les Glycines'.",
     "long", ""),
    ("1. Le nom de la maison", "1.6", "Sous-titre / accroche courte",
     "Une ligne qui décrit la maison. Ex : « Maison d'hôtes en Bourgogne du Sud ».",
     "Pour le SEO, doit contenir 'Bourgogne' ou 'Saône-et-Loire' ou 'Côte de Beaune'.",
     "text", ""),

    # ===== 2. COORDONNÉES =====
    ("2. Coordonnées", "2.1", "Adresse complète",
     "N°, voie, code postal, ville.",
     "(Pré-rempli par Morgan)",
     "text", "8 rue Jean Bouveri, 71360 Épinac"),
    ("2. Coordonnées", "2.2", "Téléphone principal",
     "Format : +33 (0)X XX XX XX XX.",
     "Privilégier un numéro mobile qui répond toujours plutôt qu'un fixe.",
     "text", ""),
    ("2. Coordonnées", "2.3", "Email de contact",
     "Idéalement contact@[nomdomaine].fr (à créer après choix du domaine).",
     "Évite Gmail/Yahoo personnels — un email pro renforce la crédibilité.",
     "text", ""),
    ("2. Coordonnées", "2.4", "GPS — Latitude",
     "(Pré-rempli — à vérifier sur Google Maps)",
     "Pré-rempli au centre d'Épinac. À affiner avec un clic droit sur Google Maps quand on aura la position exacte de la maison.",
     "text", "46.987"),
    ("2. Coordonnées", "2.5", "GPS — Longitude",
     "(Pré-rempli — à vérifier sur Google Maps)",
     "",
     "text", "4.510"),

    # ===== 3. LE BÂTIMENT & SON HISTOIRE =====
    ("3. Le bâtiment & son histoire", "3.1", "Année de construction (estimée)",
     "Année exacte ou époque (ex : 'début XIXᵉ siècle'). Si pas sûre : « À voir ensemble ».",
     "L'antériorité valorise — à mettre en avant dans le hero du site.",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.2", "Style architectural",
     "Demeure bourgeoise / Manoir / Château / Maison de maître / Ferme rénovée / Industriel / Autre.",
     "Pour positionner le site dans le bon registre.",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.3", "Nombre d'étages / niveaux",
     "Ex : RDC + 1 étage + grenier.",
     "",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.4", "Surface habitable totale",
     "En m². Approximation OK.",
     "",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.5", "Surface du parc / propriété",
     "En m². Approximation OK.",
     "Argument fort si > 5 000 m². À mettre en avant.",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.6", "Histoire courte de la maison",
     "3 à 5 phrases sur l'origine, l'usage passé.",
     "C'est ce qui différencie la maison de tous les autres B&B. Sois précise (date, événements, anciens propriétaires) — je reformule.",
     "long", ""),
    ("3. Le bâtiment & son histoire", "3.7", "Anciens propriétaires connus",
     "Noms, métiers, époques. Si tu sais.",
     "Donner des noms et des métiers fait vivre l'histoire. Ex : « famille Untel, négociants en vin, 1820-1890 ».",
     "long", ""),
    ("3. Le bâtiment & son histoire", "3.8", "Anciens usages de la maison",
     "Résidence privée / Ferme / Maison d'artisan / Exploitation viticole / Demeure de directeurs / Autre. Plusieurs usages possibles.",
     "L'usage passé nourrit le récit — particulièrement fort si lié à l'histoire industrielle d'Épinac (mines, verrerie).",
     "long", ""),
    ("3. Le bâtiment & son histoire", "3.9", "Lien avec l'histoire locale d'Épinac",
     "Lien avec la verrerie, les mines, la sidérurgie du Creusot, l'agriculture, etc.",
     "Plus le lien est fort et précis, plus le site est unique. Aucun B&B générique de Booking ne peut copier ce récit.",
     "long", ""),
    ("3. Le bâtiment & son histoire", "3.10", "La maison a-t-elle un nom historique ?",
     "Ancien nom donné par les habitants, écrit sur des cartes, des actes notariés.",
     "Argument fort pour le naming et la légitimité.",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.11", "Anecdotes, légendes ou histoires locales",
     "Une histoire transmise, un événement, une rumeur, un fantôme bienveillant…",
     "C'est le détail qui crée du bouche-à-oreille. Soigne ce qui te vient en tête.",
     "long", ""),
    ("3. Le bâtiment & son histoire", "3.12", "Statut patrimonial",
     "Classé Monument Historique / Inscrit MH / ZPPAUP / Aucun classement / Je ne sais pas.",
     "Tout classement est un argument à mettre en avant.",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.13", "Caractéristiques architecturales remarquables (intérieur)",
     "Parquet point de Hongrie, cheminées d'époque, lambris, hauteur sous plafond, escalier d'apparat, fonte Rococo, vitrail, fresques, papiers peints d'origine.",
     "Plus tu donnes de détails authentiques, plus le site sonne juste et premium.",
     "long", ""),
    ("3. Le bâtiment & son histoire", "3.14", "Matériaux extérieurs dominants",
     "Pierre locale (calcaire, granit) / Brique / Crépi / Pierre + brique / Autre.",
     "Pour décrire la façade dans le récit.",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.15", "Toiture",
     "Tuile plate / Tuile canal / Ardoise / Tuile vernissée bourguignonne / Autre.",
     "La tuile vernissée est emblématique de Bourgogne — argument fort si présente.",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.16", "Couleur dominante de la façade",
     "Pierre apparente / Crépi clair / Crépi ocre / Autre.",
     "",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.17", "Orientation principale / vue",
     "Sud / Sud-ouest / Vue sur parc / Vue sur bourg / Vue sur campagne, etc.",
     "Une belle exposition = argument à mentionner.",
     "text", ""),
    ("3. Le bâtiment & son histoire", "3.18", "Dépendances présentes sur la propriété",
     "Grange / Écurie / Tour / Pavillon / Orangerie / Chapelle / Pigeonnier / Autre. Plusieurs réponses possibles.",
     "Chaque dépendance = une histoire potentielle à raconter.",
     "long", ""),
    ("3. Le bâtiment & son histoire", "3.19", "Particularités uniques sur la propriété",
     "Four à pain ancien / Puits / Glacière / Cave voûtée / Vignes sur la propriété / Verger / Source / Pièce d'eau / Autre.",
     "Le détail rare = ce qui rend le lieu mémorable.",
     "long", ""),
    ("3. Le bâtiment & son histoire", "3.20", "Histoire de la rénovation actuelle",
     "Quand as-tu commencé ? Avec qui ? Architecte, artisans locaux, fait soi-même ? Démarche de restauration ?",
     "Le 'comment ça se rénove' est passionnant à raconter. Crée un lien émotionnel avec le futur visiteur.",
     "long", ""),
    ("3. Le bâtiment & son histoire", "3.21", "Éléments anciens conservés vs neufs",
     "Qu'est-ce qui est d'origine et qu'est-ce qui est nouveau ? Ex : parquet d'origine, cheminée d'origine + sanitaires neufs.",
     "Crédibilise le positionnement « tenue par la pierre, pas restaurée ».",
     "long", ""),

    # ===== 4. LES CHAMBRES (V1) =====
    ("4. Les chambres (V1)", "4.1", "Combien de chambres ouvres-tu pour la V1 ?",
     "On part sur 2 chambres pour septembre, c'est ça ?",
     "Si tu confirmes 2, écris '2'. Si tu hésites, dis-le.",
     "text", ""),
    ("4. Les chambres (V1)", "4.2", "Date prévue d'ouverture des chambres",
     "Mois et année (ex : septembre 2026).",
     "",
     "text", ""),
    ("4. Les chambres (V1)", "4.3", "Capacité totale en couchages (V1)",
     "Combien de personnes peuvent dormir dans les 2 chambres.",
     "",
     "text", ""),
    ("4. Les chambres (V1)", "4.4", "Petit-déjeuner inclus dans le tarif ?",
     "Oui / Non.",
     "Inclus = standard pour positionnement chaleureux haut-de-gamme.",
     "text", ""),
    ("4. Les chambres (V1)", "4.5", "Tarif moyen envisagé par nuit (€/nuit pour 2)",
     "Si tu ne sais pas encore, écris « À voir ensemble » — on cherchera des références.",
     "Pour Bourgogne du Sud avec petit-déjeuner inclus : 150-220€ standard premium.",
     "text", ""),

    # ===== 4 BIS — CHAMBRE A =====
    ("4 BIS. Chambre A", "", "🔑 PRINCIPE CLÉ : LES 2 CHAMBRES DOIVENT ÊTRE TRÈS DIFFÉRENTES",
     "Chaque chambre doit avoir son propre univers : son nom, ses couleurs dominantes, son ambiance, son caractère. Ne reproduis pas la même décoration dans les deux. Le visiteur doit avoir un VRAI choix entre 2 mondes distincts. C'est la clé du positionnement.",
     "", "info", ""),
    ("4 BIS. Chambre A", "A.1", "Nom de la Chambre A",
     "Provisoire OK.",
     "Donner un nom = personnalité. Évite « Chambre 1 » qui est froid sur le site.",
     "text", ""),
    ("4 BIS. Chambre A", "A.2", "Pour qui est-elle pensée ?",
     "Couple, voyageur seul, famille.",
     "Une chambre pensée pour un usage = description plus précise et plus engageante.",
     "text", ""),
    ("4 BIS. Chambre A", "A.3", "Surface approximative", "En m².", "", "text", ""),
    ("4 BIS. Chambre A", "A.4", "Étage", "RDC / 1er / 2e.", "RDC = accessible et + adapté aux seniors.", "text", ""),
    ("4 BIS. Chambre A", "A.5", "Atmosphère générale (3 mots max)",
     "Ex : « lumineuse, sereine, raffinée » ou « feutrée, automnale, intimiste ».",
     "Ces 3 mots vont guider la génération d'image IA. Sois précise.",
     "text", ""),
    ("4 BIS. Chambre A", "A.6", "Hauteur sous plafond approximative",
     "En mètres (estimation OK).",
     "Au-dessus de 3,5m = argument à mettre en avant dans la description.",
     "text", ""),
    ("4 BIS. Chambre A", "A.7", "Type de plafond",
     "À la française (poutres apparentes) / Plâtre simple / Avec moulures / Autre.",
     "", "choice", ""),
    ("4 BIS. Chambre A", "A.8", "Sol",
     "Parquet point de Hongrie d'origine / Parquet à l'anglaise / Tomettes / Autre.",
     "Préciser « parquet d'origine » sur le site = booste la valeur perçue.",
     "choice", ""),
    ("4 BIS. Chambre A", "A.9", "Cheminée dans la chambre ?",
     "Oui (matière : marbre / pierre / bois / faïence) / Non.",
     "Si oui, mention obligatoire dans le titre ou la description.",
     "text", ""),
    ("4 BIS. Chambre A", "A.10", "Fenêtres (nombre, taille)",
     "Ex : « deux grandes fenêtres avec petits carreaux ».",
     "Important pour la génération d'image IA.",
     "text", ""),
    ("4 BIS. Chambre A", "A.11", "Vue depuis les fenêtres",
     "Parc / Cour / Campagne / Autre.", "", "choice", ""),
    ("4 BIS. Chambre A", "A.12", "Couleur dominante des murs",
     "Ex : « vert sauge », « blanc cassé », « bleu de Prusse ».",
     "Sois aussi précise que possible — l'IA et le site en dépendent.",
     "text", ""),
    ("4 BIS. Chambre A", "A.13", "Boiseries / lambris ?",
     "Oui peintes (couleur) / Oui apparent / Non.", "", "choice", ""),
    ("4 BIS. Chambre A", "A.14", "Dimension du lit",
     "King 180 / Queen 160 / Double 140 / Lits jumeaux.", "", "choice", ""),
    ("4 BIS. Chambre A", "A.15", "Style de tête de lit",
     "Capitonnée tissu / Bois / Fer forgé / Aucune / Autre.", "", "choice", ""),
    ("4 BIS. Chambre A", "A.16", "Linge de lit",
     "Lin (couleur) / Coton blanc / Couette + plaid / Autre.",
     "« Linge de lin lavé » est un argument vendable à mettre sur le site.",
     "choice", ""),
    ("4 BIS. Chambre A", "A.17", "Mobilier principal",
     "Liste : armoire, commode, fauteuil, bureau, miroir, chevets…",
     "Si pièces anciennes, à préciser — détail signature.",
     "long", ""),
    ("4 BIS. Chambre A", "A.18", "Suspension / lustre",
     "Lustre ancien / Suspension contemporaine / Aucune / Autre.", "", "choice", ""),
    ("4 BIS. Chambre A", "A.19", "Rideaux",
     "Lin / Velours (couleur) / Tissu fleuri / Voilage / Autre.", "", "choice", ""),
    ("4 BIS. Chambre A", "A.20", "Tapis",
     "Persan / Berbère / Uni / Aucun / Autre.", "", "choice", ""),
    ("4 BIS. Chambre A", "A.21", "Objet décoratif singulier",
     "Pendule, candélabre, buste, livres anciens, trumeau…",
     "C'est l'objet qui apparaîtra sur les photos et qu'on citera dans la description.",
     "text", ""),
    ("4 BIS. Chambre A", "A.22", "Salle de bains privative ?",
     "Oui (baignoire ancienne / douche italienne / les deux) / Non.", "", "choice", ""),
    ("4 BIS. Chambre A", "A.23", "Style salle de bains",
     "Carreaux ciment ancien / Marbre / Faïence blanche / Autre.", "", "choice", ""),
    ("4 BIS. Chambre A", "A.24", "Description libre — Chambre A",
     "Paragraphe 5-10 lignes : ambiance, lumière, sensations. Comme si tu y entrais à 18h en automne.",
     "C'est cette description qui devient la base du copy de la page chambre. Écris dans tes mots, je reformule.",
     "long", ""),

    # ===== 4 BIS — CHAMBRE B =====
    ("4 BIS. Chambre B", "", "🔑 RAPPEL : la Chambre B doit avoir un univers TRÈS différent de la Chambre A",
     "Si la Chambre A est claire et fleurie, la B sera profonde et boisée. Si A est intime, B sera lumineuse. Le but : offrir 2 expériences distinctes au visiteur.",
     "", "info", ""),
    ("4 BIS. Chambre B", "B.1", "Nom de la Chambre B", "Provisoire OK.",
     "Choisir un nom complémentaire (pas redondant) à la Chambre A.",
     "text", ""),
    ("4 BIS. Chambre B", "B.2", "Pour qui est-elle pensée ?", "",
     "Différencier de la Chambre A pour offrir 2 expériences distinctes.",
     "text", ""),
    ("4 BIS. Chambre B", "B.3", "Surface approximative", "En m².", "", "text", ""),
    ("4 BIS. Chambre B", "B.4", "Étage", "RDC / 1er / 2e.", "", "text", ""),
    ("4 BIS. Chambre B", "B.5", "Atmosphère générale (3 mots max)", "",
     "Privilégier une ambiance différente de la Chambre A.",
     "text", ""),
    ("4 BIS. Chambre B", "B.6", "Hauteur sous plafond", "En mètres approx.", "", "text", ""),
    ("4 BIS. Chambre B", "B.7", "Type de plafond",
     "À la française / Plâtre simple / Avec moulures / Autre.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.8", "Sol", "Parquet / Tomettes / Autre.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.9", "Cheminée ?", "Oui (matière) / Non.", "", "text", ""),
    ("4 BIS. Chambre B", "B.10", "Fenêtres", "Nombre, taille, style.", "", "text", ""),
    ("4 BIS. Chambre B", "B.11", "Vue", "Parc / Cour / Campagne / Autre.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.12", "Couleur dominante des murs", "Description précise.", "", "text", ""),
    ("4 BIS. Chambre B", "B.13", "Boiseries / lambris ?",
     "Oui peintes / Oui apparent / Non.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.14", "Dimension du lit",
     "King / Queen / Double / Jumeaux.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.15", "Style tête de lit",
     "Capitonnée / Bois / Fer forgé / Aucune.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.16", "Linge de lit",
     "Lin / Coton / Couette / Autre.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.17", "Mobilier principal",
     "Liste : armoire, commode, fauteuil, etc.", "", "long", ""),
    ("4 BIS. Chambre B", "B.18", "Suspension / lustre", "Style.", "", "text", ""),
    ("4 BIS. Chambre B", "B.19", "Rideaux", "Lin / Velours / Fleuri / Voilage.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.20", "Tapis", "Persan / Berbère / Uni / Aucun.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.21", "Objet décoratif singulier", "Pièce unique qui marque la chambre.", "", "text", ""),
    ("4 BIS. Chambre B", "B.22", "Salle de bains privative ?", "Oui (agencement) / Non.", "", "text", ""),
    ("4 BIS. Chambre B", "B.23", "Style salle de bains",
     "Ciment / Marbre / Faïence / Autre.", "", "choice", ""),
    ("4 BIS. Chambre B", "B.24", "Description libre — Chambre B",
     "Paragraphe d'ambiance.", "", "long", ""),

    # ===== 4 BIS BONUS =====
    ("4 BIS. Bonus visuel", "BIS.1", "Palette de couleurs globale de la maison",
     "3-5 couleurs dominantes : crème, bordeaux, vert anglais, terre cuite, ocre…",
     "Une palette restreinte donne une identité visuelle forte sur le site.",
     "long", ""),
    ("4 BIS. Bonus visuel", "BIS.2", "Objets / meubles de famille à réutiliser",
     "Pièces uniques héritées, brocantées : armoire normande, piano, trumeau, malle…",
     "Le mobilier de famille = ce qui rendra la maison unique. Liste-les.",
     "long", ""),

    # ===== 5. MARIAGES (simplifié V1) =====
    ("5. Mariages (V1)", "5.1", "Tu ouvres les mariages dès la V1 ou plus tard ?",
     "Dès septembre / Plus tard (préciser quand) / À voir ensemble.",
     "Si tu préfères te concentrer sur les chambres d'abord, c'est OK.",
     "text", ""),
    ("5. Mariages (V1)", "5.2", "Capacité approximative pour cocktail debout",
     "Estimation. Si pas sûre, écris « À voir ensemble ».",
     "1 m² par personne en cocktail. À calculer en intégrant terrasse + parc.",
     "text", ""),
    ("5. Mariages (V1)", "5.3", "Capacité approximative à table assise",
     "Estimation pour un dîner.",
     "1,5 m² par personne assise.",
     "text", ""),
    ("5. Mariages (V1)", "5.4", "Saison ouverte aux mariages",
     "Ex : « avril à octobre ».",
     "",
     "text", ""),
    ("5. Mariages (V1)", "5.5", "Tarif privatisation week-end (à partir de)",
     "€ pour vendredi 14h → dimanche 14h. Si pas sûre, écris « À voir ensemble ».",
     "Pour Bourgogne du Sud avec hébergement : 4 500-8 500 €. On affinera ensemble.",
     "text", ""),
    ("5. Mariages (V1)", "5.6", "Que comprend la privatisation (en gros) ?",
     "Liste libre : domaine entier, mobilier, hébergement famille, brunch, etc.",
     "Pas besoin d'être exhaustive. Note ce qui te paraît évident, on complète.",
     "long", ""),
    ("5. Mariages (V1)", "5.7", "Combien de chambres réservées à la famille pendant un mariage ?",
     "On reste sur les 2 chambres (Chambre A pour les mariés, Chambre B pour les parents) ?",
     "Pré-rempli par Morgan. Modifie si tu vois plus large à terme.",
     "text", "2 chambres pour le moment (Chambre A pour les mariés, Chambre B pour les parents)"),

    # ===== 6. SÉMINAIRES & SALLE DE RÉCEPTION =====
    ("6. Séminaires & salle", "6.1", "Tu ouvres les séminaires dès la V1 ou plus tard ?",
     "Dès le départ / À partir de 2028 / À voir ensemble.",
     "",
     "text", ""),
    ("6. Séminaires & salle", "6.2", "Capacité estimative séminaire",
     "Nombre de personnes max envisagées.",
     "Pré-rempli par Morgan — tu peux modifier si tu veux.",
     "text", "10 à 15 personnes, voire 20 maximum"),
    ("6. Séminaires & salle", "6.3", "Tarif estimatif journée d'étude",
     "€/personne/jour. Si pas sûre, « À voir ensemble ».",
     "Standard Bourgogne pour un B&B premium : 75-120 €/pers/jour avec déjeuner. À affiner ensemble.",
     "text", ""),
    ("6. Séminaires & salle", "6.4", "🔑 SALLE PERMANENTE OU CHAPITEAU EXTÉRIEUR ?",
     "Pour les mariages et séminaires : tu envisages d'aménager une salle permanente DANS la maison ? OU tout se passe en EXTÉRIEUR sous chapiteau ?",
     "Pré-rempli par Morgan — confirme ou modifie.",
     "long", "Pour les SÉMINAIRES : idéalement, on aménage une belle salle de séminaire à l'étage de la maison, là où se trouvent les baies vitrées modernes (capacité 10-20 personnes).\n\nPour les MARIAGES : pas encore décidé. Soit on aménage un vrai espace intérieur, soit tout se passe en extérieur sur mesure, avec ou sans chapiteau de réception sur mesure.\n\n→ À discuter ensemble."),
    ("6. Séminaires & salle", "6.5", "Surface envisagée pour la salle de séminaire à l'étage ?",
     "En m² approx. À mesurer.",
     "À mesurer sur place — la salle est déjà identifiée (étage, baies vitrées modernes).",
     "text", ""),

    # ===== 7. SÉJOUR QUOTIDIEN =====
    ("7. Séjour", "7.1", "Heure d'arrivée check-in (à partir de)", "Ex : 16h.", "", "text", ""),
    ("7. Séjour", "7.2", "Heure de départ check-out", "Ex : 11h.", "", "text", ""),
    ("7. Séjour", "7.3", "Petit-déjeuner servi de…", "Ex : 8h.", "", "text", ""),
    ("7. Séjour", "7.4", "Petit-déjeuner servi à…", "Ex : 10h.", "", "text", ""),
    ("7. Séjour", "7.5", "Lieu petit-déjeuner",
     "Salle à manger / Jardin / Au choix.",
     "Au jardin l'été = signature mémorable.", "choice", ""),
    ("7. Séjour", "7.6", "Spécialités / produits proposés au petit-déjeuner",
     "Pain de Saulieu, miel de Pommard, fromages de Cîteaux…",
     "Cite les producteurs par leur nom — argument SEO + image authentique.",
     "long", ""),
    ("7. Séjour", "7.7", "Animaux acceptés ?",
     "Oui (conditions) / Non.", "", "text", ""),
    ("7. Séjour", "7.8", "Enfants accueillis ?",
     "Oui (à partir de quel âge) / Non.", "", "text", ""),

    # ===== 8. LOGO =====
    ("8. Logo", "8.1", "Note d'intention pour le logo",
     "Décris ce que tu aimerais voir dans le logo : éléments graphiques, ambiance, couleurs, références.",
     "Pas besoin de formulation parfaite. Décris en mots simples : un dessin, une initiale, une typographie élégante, un détail de la maison ?",
     "long", ""),
    ("8. Logo", "8.2", "Tu as déjà des logos / brand books inspirants ?",
     "Liens, photos, marques que tu admires.", "", "long", ""),

    # ===== 9. RÉSEAUX SOCIAUX =====
    ("9. Réseaux sociaux", "9.1", "Réseaux sociaux — note",
     "Les comptes Instagram, Facebook, Google Business Profile et LinkedIn seront créés une fois le nom du lieu validé. Rien à faire pour toi maintenant — Morgan s'en occupe.",
     "Pré-rempli — RAS de ton côté.",
     "text", "Morgan s'en occupe une fois le nom validé."),

    # ===== 10. CARNET D'ADRESSES (simplifié) =====
    ("10. Carnet d'adresses", "10.1", "Bonnes adresses découvertes depuis ton arrivée",
     "Ce que tu as trouvé de bien depuis 2 semaines : restaurants, vignerons, boulangerie, fromager, artisans, lieux.",
     "Pas besoin d'être exhaustive. Je compléterai par des recherches. Liste ce qui t'a marquée.",
     "long", ""),
    ("10. Carnet d'adresses", "10.2", "Tu as un coup de cœur particulier ?",
     "Une adresse, un événement, un personnage local que tu adores.",
     "C'est ce qui rendra le site authentique.", "long", ""),

    # ===== 11. PRÉFÉRENCES ÉDITORIALES =====
    ("11. Préférences éditoriales", "11.1", "Ce que tu veux ABSOLUMENT voir sur le site",
     "Tes priorités fortes : un sujet, un visuel, une page, un message qui doivent absolument apparaître.",
     "Ex : « Je veux qu'on parle obligatoirement du parc » / « Je veux que le mot famille apparaisse partout ».",
     "long", ""),
    ("11. Préférences éditoriales", "11.2", "Ce que tu ne veux ABSOLUMENT PAS voir sur le site",
     "Tes interdits : un sujet, un mot, une image, une pratique que tu refuses.",
     "Ex : « Pas de prix affichés sur la home » / « Pas le mot luxe nulle part » / « Pas de photo de moi ».",
     "long", ""),
    ("11. Préférences éditoriales", "11.3", "Mots / expressions à ÉVITER",
     "Le détail des mots qui te dérangent.", "", "long", ""),
    ("11. Préférences éditoriales", "11.4", "Mots / expressions à GARDER",
     "Le détail des mots que tu aimes.", "", "long", ""),
    ("11. Préférences éditoriales", "11.5", "Sites concurrents que tu aimes",
     "URL ou noms de maisons inspirantes.",
     "Indique 3-5 sites — je m'en inspirerai sans copier.",
     "long", ""),

    # ===== 12. MOT D'ACCUEIL =====
    ("12. Mot d'accueil", "12.1", "Mot de bienvenue (50-100 mots)",
     "Qui tu es, pourquoi cette maison, ce que tu y aimes, ce que tu souhaites à tes hôtes.",
     "Le détail qui change tout. Privilégier la sincérité brute, je reformulerai dans le ton du site.",
     "long", ""),

    # ===== 13. LANCEMENT & INFRASTRUCTURE =====
    ("13. Lancement & infrastructure", "13.1", "Date globale d'ouverture envisagée",
     "Mois et année.", "", "text", ""),
    ("13. Lancement & infrastructure", "13.2", "Mobilier événementiel — location ou achat ?",
     "Tu prévois d'acheter du mobilier événementiel (tables, chaises, vaisselle pour 60 personnes) ou tu préfères louer à chaque événement ?",
     "Achat = investissement initial fort + maîtrise totale. Location = pas de stockage + flexibilité. À discuter ensemble selon volume prévu.",
     "long", ""),
    ("13. Lancement & infrastructure", "13.3", "TV / écran dans les chambres ou salles communes ?",
     "Oui (préciser où) / Non / À voir.",
     "Pas de TV dans les chambres = positionnement « déconnexion » assumé. Cheminée + livres > TV pour le segment haut-de-gamme.",
     "text", ""),
    ("13. Lancement & infrastructure", "13.4", "Fibre internet installée / prévue ?",
     "Oui (date) / Pas encore (date prévue) / À voir.",
     "Indispensable pour les séminaires + qualité wifi attendue par la cible. À installer AVANT l'ouverture.",
     "text", ""),
    ("13. Lancement & infrastructure", "13.5", "Branchements électriques extérieurs prévus ?",
     "Oui (combien, où) / Pas encore / À voir.",
     "Indispensable pour les événements (sonorisation, éclairage, food trucks). Prévoir au moins 4-6 prises réparties dans le parc.",
     "long", ""),
    ("13. Lancement & infrastructure", "13.6", "Puissance électrique disponible (kVA)",
     "Si tu sais. Sinon « À voir ensemble ».",
     "Pour un événement avec sono + éclairage extérieur : 30 kVA minimum recommandé. À vérifier avec un électricien.",
     "text", ""),
    ("13. Lancement & infrastructure", "13.7", "Responsable du site (pour les mentions légales)",
     "Le nom de la personne qui apparaîtra dans la rubrique 'mentions légales' du site.",
     "Pré-rempli par Morgan. Pas besoin de modifier.",
     "text", "Frédérique Fraud"),

    # ===== 14. CHAMP LIBRE =====
    ("14. Ce que j'oublie", "14.1", "Tout ce qui mérite d'apparaître sur le site et qui n'a pas été demandé",
     "Champ libre.", "", "long", ""),
]

# ===== CONSTRUCTION DU FICHIER =====
wb = Workbook()
ws = wb.active
ws.title = "Questionnaire"

BRASS = "B59661"
BRASS_DEEP = "6B5A2E"
INK = "1C1C1C"
PAPER = "FBF8F0"
INFO_BG = "F5EFE0"
HELP_GRAY = "888888"
RULE = "DDDDDD"
PREFILL_BG = "EAF4E8"  # vert très clair pour signaler le pré-rempli

thin = Border(
    left=Side(style="thin", color=RULE),
    right=Side(style="thin", color=RULE),
    top=Side(style="thin", color=RULE),
    bottom=Side(style="thin", color=RULE),
)

headers = ["Section", "N°", "Question", "Aide / Exemple", "💡 Conseil de Morgan", "✏️ Ta réponse"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.fill = PatternFill(start_color=INK, end_color=INK, fill_type="solid")
    cell.font = Font(name="Inter", color=PAPER, bold=True, size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = thin
ws.row_dimensions[1].height = 36

widths = {"A": 26, "B": 7, "C": 42, "D": 50, "E": 55, "F": 55}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

current_section = None
row = 2
for entry in QUESTIONS:
    section, num, question, aide, conseil, qtype, prefill = entry

    if qtype == "info":
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=f"⚠️ {question}\n\n{aide}")
        cell.fill = PatternFill(start_color=INFO_BG, end_color=INFO_BG, fill_type="solid")
        cell.font = Font(name="Inter", color=BRASS_DEEP, italic=True, size=11, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        cell.border = thin
        ws.row_dimensions[row].height = 110
        row += 1
        continue

    if section != current_section:
        current_section = section
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=section)
        cell.fill = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
        cell.font = Font(name="Cormorant Garamond", italic=True, size=14, color=BRASS, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 30
        row += 1

    ws.cell(row=row, column=1, value=section).font = Font(name="Inter", color=HELP_GRAY, size=9)
    ws.cell(row=row, column=2, value=num).font = Font(name="Inter", color=BRASS, bold=True, size=10)
    ws.cell(row=row, column=3, value=question).font = Font(name="Inter", color=INK, bold=True, size=11)
    ws.cell(row=row, column=4, value=aide).font = Font(name="Inter", color=HELP_GRAY, italic=True, size=10)
    ws.cell(row=row, column=5, value=conseil).font = Font(name="Inter", color=BRASS_DEEP, italic=True, size=10)
    response_cell = ws.cell(row=row, column=6, value=prefill if prefill else "")
    response_cell.font = Font(name="Inter", color=INK, size=11, bold=True if prefill else False)

    if conseil:
        ws.cell(row=row, column=5).fill = PatternFill(start_color="FAF5E8", end_color="FAF5E8", fill_type="solid")

    if prefill:
        response_cell.fill = PatternFill(start_color=PREFILL_BG, end_color=PREFILL_BG, fill_type="solid")

    for c in range(1, 7):
        cell = ws.cell(row=row, column=c)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        cell.border = thin

    if qtype == "long":
        ws.row_dimensions[row].height = 92
    else:
        ws.row_dimensions[row].height = 50

    row += 1

ws.freeze_panes = "A2"
wb.save(out)
print(f"✓ {out.name}  ({out.stat().st_size // 1024} Ko)  · {row - 1} lignes")
